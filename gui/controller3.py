# コマンドの送信用gui
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.clock import Clock
from kivy.garden.matplotlib.backend_kivyagg import FigureCanvasKivyAgg
from kivy.core.window import Window

import openpyxl

import numpy as np
import matplotlib.pyplot as plt
import matplotlib
# Kivy 上で Matplotlib を使うために必要な準備
matplotlib.use('module://kivy.garden.matplotlib.backend_kivy')
import re
import serial
import threading

import analyseModule

# 定数
FRAME_TIME = 10 # 更新間隔(ms)
FRAME_NUM = 200 # 画面に表示するデータ数

commands = [] # 送信待ちコマンド
input_value = '' # 入力中のコマンド

# グラフ描画のstart/stop
is_graph_updating = True

# グラフのデータ
light1 = np.zeros(FRAME_NUM)
light2 = np.zeros(FRAME_NUM)
pos = np.zeros(FRAME_NUM)
time = np.zeros(FRAME_NUM)

class MainScreen3(BoxLayout):
    def handle_change(self, value):
        global input_value
        input_value = value

    def handle_submit(self, value):
        global commands
        global input_value
        commands.append(input_value)
        print(input_value)

    def handle_start(self):
        global commands
        global is_graph_updating
        commands.append("b1a")
        is_graph_updating = True
        print("start")

    def handle_stop(self):
        global commands
        global is_graph_updating
        global light1
        global light2
        global time
        global FRAME_NUM

        commands.append("b0a")
        is_graph_updating = False

        # 最初の200行は0で埋まっているので削除
        light1 = light1[201::-1]
        light2 = light2[201::-1]
        time = time[201::-1]

        # Excelにログ保存
        wb = openpyxl.load_workbook('logs/log.xlsx')
        sheet = wb['Sheet1']
        for t in range(np.size(time)):
            sheet.cell(row=t+1,column=1,value=t+1)
            sheet.cell(row=t+1,column=2,value=time[t])
            sheet.cell(row=t+1,column=3,value=light1[t])
            sheet.cell(row=t+1,column=4,value=light2[t])
        wb.save("logs/log.xlsx")

        analyseModule.analyseData(light1, light2)

        # データ初期化
        light1 =light2 = pos = time = np.zeros(FRAME_NUM)
        print("stop")

    def handle_mode_test(self):
        global commands
        commands.append("c2a")
        print("testMode")

    def handle_mode_run(self):
        global commands
        commands.append("c0a")
        print("runMode")

    def handle_kp_change(self, value):
        global commands
        commands.append("j"+str(value)+"a")
        print("KP: "+str(value))

    def handle_ki_change(self, value):
        global commands
        commands.append("k"+str(value)+"a")
        print("KI: "+str(value))

    def handle_kd_change(self, value):
        global commands
        commands.append("l"+str(value)+"a")
        print("KD: "+str(value))

class GraphView(BoxLayout):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        global FRAME_TIME
        global FRAME_NUM

        # 初期化に用いるデータ
        x = np.linspace(1,FRAME_NUM,FRAME_NUM)
        y = np.zeros(FRAME_NUM)

        # Figure, Axis を保存しておく
        self.fig, self.ax = plt.subplots(2, facecolor="0.1")
        self.ax[0].tick_params(axis='x', colors="0.8")
        self.ax[0].tick_params(axis='y', colors="0.8")
        self.ax[0].set_facecolor((0.4, 0.4, 0.4, 1))
        self.ax[1].tick_params(axis='x', colors="0.8")
        self.ax[1].tick_params(axis='y', colors="0.8")
        self.ax[1].set_facecolor((0.4, 0.4, 0.4, 1))

        # 最初に描画したときの Line も保存しておく
        self.line11, = self.ax[0].plot(x, y)
        self.line12, = self.ax[0].plot(x, y)
        self.line21, = self.ax[1].plot(x, y)

        # ウィジェットとしてグラフを追加する
        widget = FigureCanvasKivyAgg(self.fig)
        self.add_widget(widget)

        # frame_time秒ごとに表示を更新するタイマーを仕掛ける
        Clock.schedule_interval(self.update_view, FRAME_TIME/1000)

    def update_view(self, *args, **kwargs):

        global light1
        global light2
        global pos
        global is_graph_updating

        # is_graph_updating == Falseならグラフ更新しない
        if not is_graph_updating:
            return

        # データを更新する
        x = np.linspace(1,FRAME_NUM,FRAME_NUM)
        y11 = light1[-FRAME_NUM:]
        y12 = light2[-FRAME_NUM:]
        y21 = pos[-FRAME_NUM:]

        # Line にデータを設定する
        self.line11.set_data(x, y11)
        self.line12.set_data(x, y12)
        self.line21.set_data(x, y21)
        # グラフの見栄えを調整する
        self.ax[0].relim()
        self.ax[0].autoscale_view()
        self.ax[1].relim()
        self.ax[1].autoscale_view()
        # 再描画する
        self.fig.canvas.draw()
        self.fig.canvas.flush_events()

class Controller3App(App):
    def __init__(self, **kwargs):
        super(Controller3App, self).__init__(**kwargs)

    def build(self):
        serialClient = SerialClient()
        serialClient.start()
        screen = MainScreen3()
        Window.size = (1000,700)
        Window.top = 50
        return screen

# シリアル通信受信クラス
class SerialClient():
    def start(self):
        serial_thread = threading.Thread(target=self.serial_method, daemon=True)
        serial_thread.start()
    
    # シリアル通信用スレッドの実装部
    def serial_method(self):
        global commands
        ser = serial.Serial("COM5",9600) # シリアル通信

        while True:
            line = ser.readline()
            line = line.decode().rstrip('\r\n')
            receives = re.split(',', line)
            for receive in receives:
                x = re.split(':', receive)
                if x[0] == 'light1':
                    global light1
                    light1 = np.append(light1, int(x[1]))
                elif x[0] == 'light2':
                    global light2
                    light2 = np.append(light2, int(x[1]))
                elif x[0] == 'pos':
                    global pos
                    pos = np.append(pos, float(x[1]))
                elif x[0] == 'time':
                    global time
                    time = np.append(time, int(x[1]))

            # 送信
            if len(commands) > 0:
                for command in commands:
                    ser.write(command.encode())
                commands = []   

if __name__ == '__main__':
    Controller3App().run()